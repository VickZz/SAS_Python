# Loading sas modules

from sas7bdat import SAS7BDAT
import pandas as pd

# Loading SAS dataset

foo = SAS7BDAT('airline.sas7bdat')
foo.header
foo.column_names
airdata=foo.to_data_frame();
airdata
airdata['Y']
airdata['Y'][1]


# Loading Excel packages

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

# Loading Excel Workbook

wb=load_workbook('/Users/Vittorio/SAS_example/Workbook1.xlsx')

# Select first sheet

wb.get_sheet_names()
ws1 = wb.get_sheet_by_name('Sheet2')

for row in range(1, 40):
	ws1.append(range(600))

ws1['D10']=3.14
ws1['D11']=airdata['Y'][1]

wb.save(filename = '/Users/Vittorio/SAS_example/Workbook_Target.xlsx')



